import os
import re
import csv
import io
import pandas as pd
import openpyxl
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from flask import Flask, send_file

def main(velocity_names, velocity_files):

    #Create cell border style function
    def cell_border(left, right, top, bottom):
        border = Border(
            left=Side(border_style=left), 
            right=Side(border_style=right), 
            top=Side(border_style=top), 
            bottom=Side(border_style=bottom))
        return border

    #Set border style function
    def set_border(border, min_row, max_row, min_col, max_col):
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = border

    #Set number format function
    def set_format(format, min_row, max_row, min_col, max_col):
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.number_format = format

    #Create list of input files
    series_names = [re.sub('[.txt]', '', name) for name in velocity_names]

    #Create and format dataframes
    dfs = {} 
    for i, csv in enumerate(velocity_files):
        df = pd.read_csv(csv, sep='\t', header=0,
        names=['line', 'inlet_type', 'struc_from', 'struc_to', 'area', 'tc', 'intensity',
        'flow', 'velocity', 'length', 'size', 'material', 'slope'])
        df.replace({'Outfall':'OUT', 'Comb.':'COMB', 'Dp-Grate':'GRATE', 'Hdwall':'FES'}, inplace=True)
        df.replace({'Notes:  j-Line contains hyd. jump':''}, inplace=True, regex=True)
        df.replace({' j':'', '\(':'', '\)':'', ' DOUBLE':''}, inplace=True, regex=True)
        df.dropna(axis=0, inplace=True)
        for col in range(4,13):
            df.iloc[:,col] = pd.to_numeric(df.iloc[:,col], errors='coerce')
        df.drop(df.columns[0], axis=1, inplace=True)
        series = series_names[i]
        dfs[series] = df

    #Write dataframes to Excel
    output_file = io.BytesIO()
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    for series, df in dfs.items():
        df.to_excel(writer, sheet_name=series, startrow=3, startcol=1, index=False)
    writer.save()

    #Create cell border styles
    border_thin = cell_border('thin', 'thin', 'thin', 'thin')
    border_medium = cell_border('medium', 'medium', 'medium', 'medium')
    border_medium_top = cell_border('thin', 'thin', 'medium', 'thin')
    border_medium_bottom = cell_border('thin', 'thin', 'thin', 'medium')
    border_medium_left = cell_border('medium', 'thin', 'thin', 'thin')
    border_medium_right = cell_border('thin', 'medium', 'thin', 'thin')
    border_medium_tlcorner = cell_border('medium', 'thin', 'medium', 'thin')
    border_medium_trcorner = cell_border('thin', 'medium', 'medium', 'thin')
    border_medium_blcorner = cell_border('medium', 'thin', 'thin', 'medium')
    border_medium_brcorner = cell_border('thin', 'medium', 'thin', 'medium')
    border_medium_top_bottom = cell_border('thin', 'thin', 'medium', 'medium')
    border_medium_top_left_bottom = cell_border('medium', 'thin', 'medium', 'medium')
    border_medium_top_right_bottom = cell_border('thin', 'medium', 'medium', 'medium')

    #Format Excel Sheets
    wb = load_workbook(output_file)

    for i, ws in enumerate(wb.worksheets):

        row_count = ws.max_row

        ws['B2'] = series_names[i] + ' SERIES (2-YEAR ANALYSIS)'
        ws['B3'] = 'INLET TYPE'
        ws['C3'] = 'STRUCTURE'
        ws['E3'] = 'A (TOTAL)'
        ws['F3'] = 'TC'
        ws['G3'] = 'I'
        ws['H3'] = 'Q'
        ws['I3'] = 'V'
        ws['J3'] = 'PIPE LENGTH'
        ws['K3'] = 'PIPE SIZE'
        ws['L3'] = 'MATERIAL'
        ws['M3'] = 'SLOPE'

        ws['B4'] = ''
        ws['C4'] = 'FROM'
        ws['D4'] = 'TO'
        ws['E4'] = '(AC)'
        ws['F4'] = '(MIN)'
        ws['G4'] = '(IN/HR)'
        ws['H4'] = '(CFS)'
        ws['I4'] = '(FT/S)'
        ws['J4'] = '(FT)'
        ws['K4'] = '(IN)'
        ws['L4'] = ''
        ws['M4'] = '(%)'

        #Replace line number with structure name
        structures = []
        for row in ws.iter_rows(min_row=5, max_row=row_count, min_col=3, max_col=3):
            for cell in row:
                structures.append(cell.value)

        for row in ws.iter_rows(min_row=5, max_row=row_count, min_col=4, max_col=4):
            for cell in row:
                if cell.value == 'OUT':
                    pass
                else:
                    line = int(cell.value)
                    cell.value = structures[line-1]

        #Replace n value with material type
        for row in ws.iter_rows(min_row=5, max_row=row_count, min_col=12, max_col=12):
            for cell in row:
                cell.value = 'RCP'

        #Merge cells
        ws.merge_cells('B2:M2')
        ws.merge_cells('C3:D3')

        #Set cell font
        for row in ws.iter_rows():
            for cell in row:      
                cell.alignment =  Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=10, name='Arial')

        for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=13):
            for cell in row:
                cell.font = Font(bold=True, size=10, name='Arial')

        #Set cell background fill
        ws['B2'].fill = PatternFill(fgColor='BFBFBF', bgColor='BFBFBF', fill_type = 'solid')

        for row in ws.iter_rows(min_row=3, max_row=4, min_col=2, max_col=13):
            for cell in row:
                cell.fill = PatternFill(fgColor='D9D9D9', bgColor='D9D9D9', fill_type = 'solid')

        #Set cell border
        set_border(border_thin, 3, row_count, 2, 13)
        set_border(border_medium_top_left_bottom, 2, 2, 2, 2)
        set_border(border_medium_top_right_bottom, 2, 2, 13, 13)
        set_border(border_medium_tlcorner, 3, 3, 2, 2)
        set_border(border_medium_left, 4, 4, 2, 2)
        set_border(border_medium_trcorner, 3, 3, 13, 13)
        set_border(border_medium_right, 4, 4, 13, 13)
        set_border(border_medium_tlcorner, 5, 5, 2, 2)
        set_border(border_medium_trcorner, 5, 5, 13, 13)
        set_border(border_medium_top_bottom, 2, 2, 3, 12)
        set_border(border_medium_top, 5, 5, 3, 12)
        set_border(border_medium_left, 6, row_count, 2, 2)
        set_border(border_medium_right, 6, row_count, 13, 13)
        set_border(border_medium_bottom, row_count, row_count, 3, 12)
        set_border(border_medium_blcorner, row_count, row_count, 2, 2)
        set_border(border_medium_brcorner, row_count, row_count, 13, 13)

        #Set cell number format
        set_format('0.00', 5, row_count, 5, 5)
        set_format('0.0', 5, row_count, 6, 6)
        set_format('0.00', 5, row_count, 7, 9)
        set_format('0', 5, row_count, 10, 11)
        set_format('0.00', 5, row_count, 13, 13)

        #Set row and column dimensions
        ws.row_dimensions[1].height = 13.8
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 21
        ws.row_dimensions[4].height = 21

        ws.column_dimensions['A'].width = 4.02
        ws.column_dimensions['B'].width = 13.13
        ws.column_dimensions['C'].width = 10.02
        ws.column_dimensions['D'].width = 10.02
        ws.column_dimensions['E'].width = 12.30
        ws.column_dimensions['F'].width = 9.58
        ws.column_dimensions['G'].width = 9.58
        ws.column_dimensions['H'].width = 9.58
        ws.column_dimensions['I'].width = 9.58
        ws.column_dimensions['J'].width = 15.58
        ws.column_dimensions['K'].width = 11.47
        ws.column_dimensions['L'].width = 13.91
        ws.column_dimensions['M'].width = 11.80

    wb.save(output_file)
    output_file.seek(0)
    return send_file(output_file, attachment_filename='Pipe Velocity.xlsx', as_attachment=True)

if __name__ == '__main__':    
    main()
