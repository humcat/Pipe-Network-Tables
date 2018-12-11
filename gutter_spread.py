import os
import re
import csv
import io
import pandas as pd
import openpyxl
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from flask import Flask, send_file

def main(spread_names, spread_files):

    #Create cell border style function
    def cell_border(left, right, top, bottom):
        border = Border(
            left=Side(border_style=left), 
            right=Side(border_style=right), 
            top=Side(border_style=top), 
            bottom=Side(border_style=bottom))
        return border

    #Set border style function
    def set_border(border, min_row, max_row, min_col, max_col):
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = border

    #Set number format function
    def set_format(format, min_row, max_row, min_col, max_col):
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.number_format = format

    #Create list of input files
    series_names = [re.sub('[.txt]', '', name) for name in spread_names]  

    #Create and format dataframes
    dfs = {} 
    for i, csv in enumerate(spread_files):
        df = pd.read_csv(csv, sep='\t', header=0,
        names=['line', 'structure', 'inlet_type', 'bypass', 'area', 'tc', 'intensity', 'c_value',
        'flow_inlet', 'flow_bypass', 'flow_captured', 'flow_bypassed', 'slope', 'spread'])
        df.replace({'Comb.':'COMB', 'Dp-Grate':'GRATE', 'Hdwall':'FES', 'Offsite':'NONE', 'Sag':'SAG',}, inplace=True)
        df.replace({'Notes:  j-Line contains hyd. jump':''}, inplace=True, regex=True)
        df.replace({' j':'', '\(':'', '\)':'', ' DOUBLE':''}, inplace=True, regex=True)
        df.dropna(axis=0, inplace=True)
        for col in range(4,14):
            df.iloc[:,col] = pd.to_numeric(df.iloc[:,col], errors='coerce')
        df.drop(df.columns[0], axis=1, inplace=True)
        df.loc[df.bypass == 'SAG', 'slope'] ='SAG'
        df.loc[df.inlet_type == 'GRATE', 'slope'] ='N/A'
        df.loc[df.inlet_type == 'GRATE', 'spread'] ='N/A'
        series = series_names[i]
        dfs[series] = df

    #Write dataframes to Excel
    output_file = io.BytesIO()
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    for series, df in dfs.items():
        df.to_excel(writer, sheet_name=series, startrow=3, startcol=1, index=False)
    writer.save()    

    #Create cell border styles
    border_thin = cell_border('thin', 'thin', 'thin', 'thin')
    border_medium = cell_border('medium', 'medium', 'medium', 'medium')
    border_medium_top = cell_border('thin', 'thin', 'medium', 'thin')
    border_medium_bottom = cell_border('thin', 'thin', 'thin', 'medium')
    border_medium_left = cell_border('medium', 'thin', 'thin', 'thin')
    border_medium_right = cell_border('thin', 'medium', 'thin', 'thin')
    border_medium_tlcorner = cell_border('medium', 'thin', 'medium', 'thin')
    border_medium_trcorner = cell_border('thin', 'medium', 'medium', 'thin')
    border_medium_blcorner = cell_border('medium', 'thin', 'thin', 'medium')
    border_medium_brcorner = cell_border('thin', 'medium', 'thin', 'medium')
    border_medium_top_bottom = cell_border('thin', 'thin', 'medium', 'medium')
    border_medium_top_left_bottom = cell_border('medium', 'thin', 'medium', 'medium')
    border_medium_top_right_bottom = cell_border('thin', 'medium', 'medium', 'medium')

    #Format Excel Sheets
    wb = load_workbook(output_file)
    
    for i, ws in enumerate(wb.worksheets):

        row_count = ws.max_row

        ws['B2'] = series_names[i] + ' SERIES'
        ws['B3'] = 'STRUCTURE'
        ws['C3'] = 'INLET\nTYPE'
        ws['D3'] = 'BYPASS\nSTRUCTURE'
        ws['E3'] = 'DRAINAGE\nAREA'
        ws['F3'] = 'TC'
        ws['G3'] = 'I'
        ws['H3'] = 'C'
        ws['I3'] = 'Q\n(INLET)'
        ws['J3'] = 'Q\n(BYPASS)'
        ws['K3'] = 'Q\n(CAPTURED)'
        ws['L3'] = 'Q\n(BYPASSED)'
        ws['M3'] = 'LONG\nSLOPE'
        ws['N3'] = 'GUTTER\nSPREAD'

        ws['B4'] = ''
        ws['C4'] = ''
        ws['D4'] = ''
        ws['E4'] = '(AC)'
        ws['F4'] = '(MIN)'
        ws['G4'] = '(IN/HR)'
        ws['H4'] = ''
        ws['I4'] = '(CFS)'
        ws['J4'] = '(CFS)'
        ws['K4'] = '(CFS)'
        ws['L4'] = '(CFS)'
        ws['M4'] = '(FT/FT)'
        ws['N4'] = '(FT)'

        #Merge cells
        ws.merge_cells('B2:N2')

        #Set cell font
        for row in ws.iter_rows():
            for cell in row:      
                cell.alignment =  Alignment(horizontal='center', vertical='center', wrapText=True)
                cell.font = Font(size=10, name='Arial')

        for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=14):
            for cell in row:
                cell.font = Font(bold=True, size=10, name='Arial')

        #Set cell background fill
        ws['B2'].fill = PatternFill(fgColor='BFBFBF', bgColor='BFBFBF', fill_type = 'solid')

        for row in ws.iter_rows(min_row=3, max_row=4, min_col=2, max_col=14):
            for cell in row:
                cell.fill = PatternFill(fgColor='D9D9D9', bgColor='D9D9D9', fill_type = 'solid')

        #Set cell border
        set_border(border_thin, 3, row_count, 2, 14)
        set_border(border_medium_top_left_bottom, 2, 2, 2, 2)
        set_border(border_medium_top_right_bottom, 2, 2, 14, 14)
        set_border(border_medium_tlcorner, 3, 3, 2, 2)
        set_border(border_medium_left, 4, 4, 2, 2)
        set_border(border_medium_trcorner, 3, 3, 14, 14)
        set_border(border_medium_right, 4, 4, 14, 14)
        set_border(border_medium_tlcorner, 5, 5, 2, 2)
        set_border(border_medium_trcorner, 5, 5, 14, 14)

        set_border(border_medium_top_bottom, 2, 2, 3, 13)
        set_border(border_medium_top, 5, 5, 3, 13)
        set_border(border_medium_left, 6, row_count, 2, 2)
        set_border(border_medium_right, 6, row_count, 14, 14)
        set_border(border_medium_bottom, row_count, row_count, 3, 13)
        set_border(border_medium_blcorner, row_count, row_count, 2, 2)
        set_border(border_medium_brcorner, row_count, row_count, 14, 14)

        #Set cell number format
        set_format('0.00', 5, row_count, 5, 5)
        set_format('0.0', 5, row_count, 6, 6)
        set_format('0.00', 5, row_count, 7, 12)
        set_format('0.00', 5, row_count, 13, 14)

        #Set row and column dimensions
        ws.row_dimensions[1].height = 13.8
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 36
        ws.row_dimensions[4].height = 21

        ws.column_dimensions['A'].width = 4.02
        ws.column_dimensions['B'].width = 13.98
        ws.column_dimensions['C'].width = 11.24
        ws.column_dimensions['D'].width = 15.36
        ws.column_dimensions['E'].width = 13.24
        ws.column_dimensions['F'].width = 11.13
        ws.column_dimensions['G'].width = 11.13
        ws.column_dimensions['H'].width = 11.13
        ws.column_dimensions['I'].width = 13.47
        ws.column_dimensions['J'].width = 13.47
        ws.column_dimensions['K'].width = 13.47
        ws.column_dimensions['L'].width = 13.47
        ws.column_dimensions['M'].width = 13.47
        ws.column_dimensions['N'].width = 13.47

    wb.save(output_file)
    output_file.seek(0)
    return send_file(output_file, attachment_filename='Pipe Design.xlsx', as_attachment=True)

if __name__ == '__main__':    
    main()
