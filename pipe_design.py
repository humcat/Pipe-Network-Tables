import os
import re
import csv
import io
import pandas as pd
import openpyxl
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from flask import Flask, send_file

def main(design_names, design_files):

    #Create cell border style function
    def cell_border(left, right, top, bottom):
        border = Border(
            left=Side(border_style=left), 
            right=Side(border_style=right), 
            top=Side(border_style=top), 
            bottom=Side(border_style=bottom))
        return border

    #Set border style function
    def set_border(border, min_row, max_row, min_col, max_col):
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = border

    #Set number format function
    def set_format(format, min_row, max_row, min_col, max_col):
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.number_format = format    
                
    #Create list of input files
    series_names = [re.sub('[.txt]', '', name) for name in design_names]         

    #Create and format dataframes
    dfs = {} 
    for i, csv in enumerate(design_files):
        df = pd.read_csv(csv, sep='\t', header=0,
        names=['line', 'inlet_type', 'struc_from', 'struc_to', 'area', 'tc', 'intensity', 'c_value',
        'flow_inlet', 'flow_total', 'flow_cap', 'length', 'size', 'material', 'slope', 'inv_up',
        'inv_down', 'rim_up', 'rim_down', 'hgl_up', 'hgl_down'])
        df.replace({'Outfall':'OUT', 'Comb.':'COMB', 'Dp-Grate':'GRATE', 'Hdwall':'FES'}, inplace=True)
        df.replace({'Notes:  j-Line contains hyd. jump':''}, inplace=True, regex=True)
        df.replace({' j':'', '\(':'', '\)':'', ' DOUBLE':''}, inplace=True, regex=True)
        df.dropna(axis=0, inplace=True)
        for col in range(4,21):
            df.iloc[:,col] = pd.to_numeric(df.iloc[:,col], errors='coerce')
        df.drop(df.columns[0], axis=1, inplace=True)
        series = series_names[i]
        dfs[series] = df 
        
    #Write dataframes to Excel
    output_file = io.BytesIO()
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    for series, df in dfs.items():
        df.to_excel(writer, sheet_name=series, startrow=3, startcol=1, index=False)
    writer.save()

    #Create cell border styles
    border_thin = cell_border('thin', 'thin', 'thin', 'thin')
    border_medium = cell_border('medium', 'medium', 'medium', 'medium')
    border_medium_top = cell_border('thin', 'thin', 'medium', 'thin')
    border_medium_bottom = cell_border('thin', 'thin', 'thin', 'medium')
    border_medium_left = cell_border('medium', 'thin', 'thin', 'thin')
    border_medium_right = cell_border('thin', 'medium', 'thin', 'thin')
    border_medium_tlcorner = cell_border('medium', 'thin', 'medium', 'thin')
    border_medium_trcorner = cell_border('thin', 'medium', 'medium', 'thin')
    border_medium_blcorner = cell_border('medium', 'thin', 'thin', 'medium')
    border_medium_brcorner = cell_border('thin', 'medium', 'thin', 'medium')
    border_medium_top_bottom = cell_border('thin', 'thin', 'medium', 'medium')
    border_medium_top_left_bottom = cell_border('medium', 'thin', 'medium', 'medium')
    border_medium_top_right_bottom = cell_border('thin', 'medium', 'medium', 'medium')

    #Format Excel Sheets
    wb = load_workbook(output_file)

    for i, ws in enumerate(wb.worksheets):

        row_count = ws.max_row

        ws['B2'] = series_names[i] + ' SERIES (10-YEAR ANALYSIS)'
        ws['B3'] = 'INLET TYPE'
        ws['C3'] = 'STRUCTURE'
        ws['E3'] = 'A'
        ws['F3'] = 'TC'
        ws['G3'] = 'I'
        ws['H3'] = 'C'
        ws['I3'] = 'Q (INLET)'
        ws['J3'] = 'Q (TOTAL)'
        ws['K3'] = 'Q (CAPACITY)'
        ws['L3'] = 'PIPE LENGTH'
        ws['M3'] = 'PIPE SIZE'
        ws['N3'] = 'MATERIAL'
        ws['O3'] = 'PIPE SLOPE'
        ws['P3'] = 'UPPER INV'
        ws['Q3'] = 'LOWER INV'
        ws['R3'] = 'RIM ELEV UP'
        ws['S3'] = 'RIM ELEV DOWN'
        ws['T3'] = 'HGL UP'
        ws['U3'] = 'HGL DOWN'

        ws['B4'] = ''
        ws['C4'] = 'FROM'
        ws['D4'] = 'TO'
        ws['E4'] = '(AC)'
        ws['F4'] = '(MIN)'
        ws['G4'] = '(IN/HR)'
        ws['H4'] = ''
        ws['I4'] = '(CFS)'
        ws['J4'] = '(CFS)'
        ws['K4'] = '(CFS)'
        ws['L4'] = '(FT)'
        ws['M4'] = '(IN)'
        ws['N4'] = ''
        ws['O4'] = '(%)'
        ws['P4'] = '(FT)'
        ws['Q4'] = '(FT)'
        ws['R4'] = '(FT)'
        ws['S4'] = '(FT)'
        ws['T4'] = '(FT)'
        ws['U4'] = '(FT)'

        #Replace line number with structure name
        structures = []
        for row in ws.iter_rows(min_row=5, max_row=row_count, min_col=3, max_col=3):
            for cell in row:
                structures.append(cell.value)

        for row in ws.iter_rows(min_row=5, max_row=row_count, min_col=4, max_col=4):
            for cell in row:
                if cell.value == 'OUT':
                    pass
                else:
                    line = int(cell.value)
                    cell.value = structures[line-1]

        #Replace n value with material type
        for row in ws.iter_rows(min_row=5, max_row=row_count, min_col=14, max_col=14):
            for cell in row:
                cell.value = 'RCP'

        #Merge cells
        ws.merge_cells('B2:U2')
        ws.merge_cells('C3:D3')

        #Set cell font
        for row in ws.iter_rows():
            for cell in row:      
                cell.alignment =  Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=10, name='Arial')

        for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=21):
            for cell in row:
                cell.font = Font(bold=True, size=10, name='Arial')

        #Set cell background fill
        ws['B2'].fill = PatternFill(fgColor='BFBFBF', bgColor='BFBFBF', fill_type = 'solid')

        for row in ws.iter_rows(min_row=3, max_row=4, min_col=2, max_col=21):
            for cell in row:
                cell.fill = PatternFill(fgColor='D9D9D9', bgColor='D9D9D9', fill_type = 'solid')

        #Set cell border
        set_border(border_thin, 3, row_count, 2, 21)
        set_border(border_medium_top_left_bottom, 2, 2, 2, 2)
        set_border(border_medium_top_right_bottom, 2, 2, 21, 21)
        set_border(border_medium_tlcorner, 3, 3, 2, 2)
        set_border(border_medium_left, 4, 4, 2, 2)
        set_border(border_medium_trcorner, 3, 3, 21, 21)
        set_border(border_medium_right, 4, 4, 21, 21)
        set_border(border_medium_tlcorner, 5, 5, 2, 2)
        set_border(border_medium_trcorner, 5, 5, 21, 21)
        set_border(border_medium_top_bottom, 2, 2, 3, 20)
        set_border(border_medium_top, 5, 5, 3, 20)
        set_border(border_medium_left, 6, row_count, 2, 2)
        set_border(border_medium_right, 6, row_count, 21, 21)
        set_border(border_medium_bottom, row_count, row_count, 3, 20)
        set_border(border_medium_blcorner, row_count, row_count, 2, 2)
        set_border(border_medium_brcorner, row_count, row_count, 21, 21)

        #Set cell number format
        set_format('0.00', 5, row_count, 5, 5)
        set_format('0.0', 5, row_count, 6, 6)
        set_format('0.00', 5, row_count, 7, 11)
        set_format('0', 5, row_count, 12, 13)
        set_format('0.00', 5, row_count, 15, 21)

        #Set row and column dimensions
        ws.row_dimensions[1].height = 13.8
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 21
        ws.row_dimensions[4].height = 21

        ws.column_dimensions['A'].width = 4.02
        ws.column_dimensions['B'].width = 13.13
        ws.column_dimensions['C'].width = 10.02
        ws.column_dimensions['D'].width = 10.02
        ws.column_dimensions['E'].width = 9.58
        ws.column_dimensions['F'].width = 9.58
        ws.column_dimensions['G'].width = 9.58
        ws.column_dimensions['H'].width = 9.58
        ws.column_dimensions['I'].width = 12.24
        ws.column_dimensions['J'].width = 12.24
        ws.column_dimensions['K'].width = 15.24
        ws.column_dimensions['L'].width = 15.58
        ws.column_dimensions['M'].width = 11.47
        ws.column_dimensions['N'].width = 13.91
        ws.column_dimensions['O'].width = 14.24
        ws.column_dimensions['P'].width = 15.13
        ws.column_dimensions['Q'].width = 15.13
        ws.column_dimensions['R'].width = 17.24
        ws.column_dimensions['S'].width = 17.24
        ws.column_dimensions['T'].width = 13.8
        ws.column_dimensions['U'].width = 13.8

    wb.save(output_file)
    output_file.seek(0)
    return send_file(output_file, attachment_filename='Pipe Design.xlsx', as_attachment=True)

if __name__ == '__main__':    
    main()
